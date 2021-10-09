from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border,Font, Alignment,Side
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import warnings
import datetime
import time
import os
import sys
import logging
from logging.handlers import RotatingFileHandler

def cls():
    os.system('cls' if os.name=='nt' else 'clear')
    
def AddSheet(ws,week,POC):
    ws.append(['Shift','Engineer Name','Case','Type','Case','Type','Case','Type','Case','Type'])
    x=week.keys()
    code = 0
    for i in x:
        if i in POC or "POC" in week[i]:
            code+=1
            ws.append([week[i],i])
            ws.append(["",""])
        else:
            ws.append([week[i],i])

    OrangeFill = PatternFill(start_color='FFC000',end_color='FFC000',fill_type='solid')
    BlueFill = PatternFill(start_color='00B0F0',end_color='00B0F0',fill_type='solid')
    BlackFill = PatternFill(start_color='000000',end_color='000000',fill_type='solid')
    WhiteFill = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')
    GreenFill = PatternFill(start_color='33CC33',end_color='33CC33',fill_type='solid')
	
    ThinBorder = Border(left=Side(style='thick'),right=Side(style='thick'),top=Side(style='thick'),bottom=Side(style='thick'))

    BlackFont= Font(name='Cambria',size=11,bold=True,italic=True)
    WhiteFont = Font(name='Cambria',size=11,bold=True,italic=True,color='FFFFFF')

    Al= Alignment(horizontal='center',vertical='center')

    for col in range(1, ws.max_column + 1):
        cell_header = ws.cell(1, col)
        cell_header.fill = GreenFill
        cell_header.border = ThinBorder
        cell_header.font = BlackFont
        cell_header.alignment = Al
    for cell in ws['A2:A{}'.format(ws.max_row)]:
        cell[0].fill = OrangeFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['B2:B{}'.format(ws.max_row)]:
        cell[0].fill = OrangeFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['D2:D{}'.format(ws.max_row)]: 
        cell[0].fill = BlueFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['F2:F{}'.format(ws.max_row)]: 
        cell[0].fill = BlueFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['H2:H{}'.format(ws.max_row)]: 
        cell[0].fill = BlueFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['J2:J{}'.format(ws.max_row)]:
        cell[0].fill = BlueFill
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['I2:I{}'.format(ws.max_row)]:
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['C2:C{}'.format(ws.max_row)]:
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['E2:E{}'.format(ws.max_row)]:
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
    for cell in ws['G2:G{}'.format(ws.max_row)]:
        cell[0].border = ThinBorder
        cell[0].font = BlackFont
        cell[0].alignment = Al
        
    for row in range(1, ws.max_row + 1):
        cell_header = ws.cell(row, 1)
        if "POC" in cell_header.value:
            cell_header.fill = BlackFill
            cell_header.font = WhiteFont
            cell_header = ws.cell(row, 2)
            cell_header.fill = BlackFill
            cell_header.font = WhiteFont 
        if "" == cell_header.value:
            for cell in ws['A{}:J{}'.format(row,row)]:
                for ele in cell:
                    ele.fill = WhiteFill

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 26           
    ws.column_dimensions['C'].width = 12            
    ws.column_dimensions['D'].width = 12            
    ws.column_dimensions['E'].width = 12            
    ws.column_dimensions['F'].width = 12            
    ws.column_dimensions['G'].width = 12            
    ws.column_dimensions['H'].width = 12            
    ws.column_dimensions['I'].width = 12            
    ws.column_dimensions['J'].width = 12
    return code

def ExtractData(file):
    wb = load_workbook(file, data_only = True)
    logging.info("Successfully opened the excel file")
    ws = wb.active
    sh = wb[ws.title]
    temp=[]
    for row in sh.rows:
        tempr=[]
        for cell in row:
            if cell.value=='A - (6:00 AM - 3:00 PM)':
                logging.info("Successfully extracted data from the file")
                return temp[1:-1]
            tempr.append(cell.value)
        temp.append(tempr)
        
def roster(flag,temp):
    week=[]
    if flag <= 27:
        x=5
    else:
        x=32-flag
    for ele in range(x):
        week1={}
        for i in temp[2:]:
            shifts = i[flag+ele]
            if shifts in ['L', None ,'HK','TR']:
                continue
            if shifts == "B":
                shifts='G'
            week1[i[0]] = shifts 
        week.append({k: v for k, v in sorted(week1.items(), key=lambda item: item[1])})
    return week

def createExcel(week,flag,fname):
    wb = Workbook()
    POC=[]
    
    if flag <= 27:
        x=5
    else:
        x=32-flag
        
    code = []
    for i in range(x):
        ws = wb.create_sheet(str(flag+i), i)
        code.append(AddSheet(ws,week[i],POC))
    d = time.strptime(fname, "%d %b %Y")  
    weekNumber = datetime.date(d.tm_year,d.tm_mon,d.tm_mday).strftime("%V")
    if min(code)<4:
        logging.warning(str(min(code))+" POCs were found!... please check if their names were mis-spelled :"+" ".join(POC)+" week"+weekNumber)
    logging.info("Successfully Identified the Date on Monday for week"+weekNumber+" Date: "+str(flag)+" "+c[0][:3]+" "+c[1])
    try:
        wb.save('Output/Week-'+weekNumber+'.xlsx')
        logging.info("Successfully saved the excel file")
    except PermissionError:
        logging.error("Unable to open the excel file!!.. Please check if the file is opened by another program already. If so, please close and retry!")
        print("Please check if the file is opened by another program already")
        print("Due to file being already open elsewhere process terminated!!!... Please close the file and retry.")
        sys.exit()
    
if __name__ == "__main__":
    fmtstr = "%(asctime)s | %(levelname)s | Line:%(lineno)s | %(funcName)s : %(message)s"
    datestr = "%d-%m-%Y %I:%M:%S %p"
	
    logging.basicConfig(handlers=[RotatingFileHandler('Logs\output.log', maxBytes=1000000, backupCount=10)],level = logging.DEBUG,format = fmtstr,datefmt = datestr)
    warnings.simplefilter("ignore")
    c = [ele for ele in os.listdir() if ".xlsx" in ele]
    file = c[0]
    tries = 1
    while not os.path.isfile(file) and tries <5:          
        logging.warning("Unable to locate the entered file in the current directory!!... The file "+file+"NOT Found")        
        cls()
        print("*Error: entered file name should be in the month-year format eg. August-2021 \nand should exist in the same path as the .exe file!...\n")
        c = [ele for ele in os.listdir() if ".xlsx" in ele]
        file = c[0]
        tries += 1
        if tries == 5:
            logging.error("Unable to locate the entered file in the current directory!!... The file "+file+"NOT Found")
            cls()
            print("You've exhausted your 5 retry counts!!... was unable to find a file named : "+file)
            exi=input("\n Press any key to exit...  ")
            sys.exit()
    val = ExtractData(file)
    c = file.split("-")
    temp=[]
    for i in val:
        temp.append(i[1:33])
    for i in range(len(temp[0])):
        if temp[0][i]=='Mon':
            flag=i
            break
    while flag<32:
        fname = str(flag)+" "+c[0][:3]+" "+c[1][:-5]
        week1=roster(flag,temp)
        createExcel(week1,flag,fname)
        flag+=7
     
    print("""    
    
Please collect your files from the output folder.
Thank you for using this code.
This code was built by Naman Gupta.
Please feel free to leave a comment, grievance or suggestion on: guptanaman0555@gmail.com
Do follow me on GitHub at: https://github.com/GuptaNaman1998
           
    """)
    exi=input("Press any key to exit...  ")